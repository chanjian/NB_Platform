from django.shortcuts import render, redirect

from utils.link import filter_reverse
from web import models
from utils.pager import Pagination
from django import forms
from utils.bootstrap import BootStrapForm
from utils.response import BaseResponse
from django.http import JsonResponse, HttpResponse


def policy_list(request):
    queryset = models.PricePolicy.objects.all().order_by('count')
    pager = Pagination(request, queryset)
    return render(request, 'policy_list.html', {'pager': pager})


class PolicyModelForm(BootStrapForm, forms.ModelForm):
    class Meta:
        model = models.PricePolicy
        fields = "__all__"


def policy_add(request):
    if request.method == "GET":
        form = PolicyModelForm()
        return render(request, 'form4.html', {'form': form})
    form = PolicyModelForm(data=request.POST)
    if not form.is_valid():
        return render(request, 'form4.html', {'form': form})
    form.save()
    return redirect('/policy/list/')


def policy_edit(request, pk):
    instance = models.PricePolicy.objects.filter(id=pk).first()
    if request.method == "GET":
        form = PolicyModelForm(instance=instance)
        return render(request, 'form4.html', {'form': form})
    form = PolicyModelForm(data=request.POST, instance=instance)
    if not form.is_valid():
        return render(request, 'form4.html', {'form': form})
    form.save()
    return redirect(filter_reverse(request,'/policy/list/'))


def policy_delete(request):
    res = BaseResponse(status=True)
    cid = request.GET.get('cid')
    models.PricePolicy.objects.filter(id=cid).delete()
    return JsonResponse(res.dict)

def policy_upload(request):
    """批量上传【基于excel】"""
    from openpyxl import load_workbook
    #1.获取用户上传的文件对象
    file_object = request.FILES.get('exc')
    print(file_object)

    #2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object,data_only=True)
    sheet = wb.worksheets[0]

    #第一行第一列
    # cell = sheet.cell(1,1)
    # print(cell.value)

    #3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        print(row)
        print(row[0].value,row[2].value)
        qb_num = row[0].value
        qb_price = row[2].value

        exists = models.PricePolicy.objects.filter(count=qb_num).exists()
        if not exists:
            models.PricePolicy.objects.create(count=qb_num,price=qb_price)


    return redirect('policy_list')